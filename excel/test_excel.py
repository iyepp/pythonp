import requests as req
import bs4
import openpyxl as xl
import pandas as pd

# csv������ �ҷ�����
df = pd.read_csv('./The list.csv')

links = df['Link']
#print(links)

# �о� �� �������� �ּ�
url = links

# �������� ���� �о� ����
for u in url:
    res = req.get(u)

# �о� �� �������� ���� ���

# Ư�� ��Ҹ� �˻��ϱ�, �˻��� ��� ����ϱ�
    bs = bs4.BeautifulSoup(res.text, features = "html.parser") # ������ �� html �м��ϱ�

    info = bs.select(".productnameandprice-container-standard")
    details = bs.select("#accordion-product-details")
    infoList = [] # ����Ʈ�� ����/ ��ǰ���� (��)���

    for i in info:
        name = i.select_one('h1.product-name.product-detail-product-name').getText().strip()
        price = i.select_one('div.product-price.product-detail-price').getText().strip()
    #print(name)
    #print(price)

    for d in details:
        pn = d.select_one('div.style-number-title').span.getText().strip()
        img = "http:" + d.select_one('picture.product-thumb').img['srcset']
    #print(pn)
    #print('https:' + img)

        infoList.append({'name':name, 'price':price, 'pn':pn, 'img':img}) # ���� ��� ���� ����Ʈ�� ���Ѵ�/ ��ǰ���� �߰�
        #print(infoList)

# �ҷ��� ���� ������ �����ϱ�

    wb = xl.Workbook() #������ ����
    sheet = wb.active  #������ �⺻ ��Ʈ�� ������ �´�
    sheet.title = "GUCCI LIST"

    row = 1 #���� �ι�° ������ ����
    col = 1 #���� ù��° �࿡�� ����

    for c in infoList:
        sheet.cell(row = row, column = col).value = c["img"]
        col = col + 2
        for c in infoList:
            sheet.cell(row = row + 1, column = col).value = c["name"]
            sheet.cell(row = row + 3, column = col).value = c["pn"]
            col = col + 1
            for c in infoList:
                sheet.cell(row = row + 4, column = col).value = c["price"]
                row = row + 10

    wb.save("DS with Python.xlsx")




