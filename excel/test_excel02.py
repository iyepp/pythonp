# -*- conding: utf-8 -*-
import requests as req
import bs4
import openpyxl as xl
import pandas as pd

infoList=[]
for _ in range(2):
    cmd = input("cmd : ")
    ret = input("ret : ")
    infoList.append({'cmd':cmd, 'ret':ret})

wb = xl.Workbook()
sheet = wb.active 
sheet.title = "GUCCI LIST"

row = 1
col = 1

for c in infoList:
        sheet.cell(row = row , column = col+1).value = c["cmd"]
        sheet.cell(row = row , column = col+3).value = c["ret"]
        row= row + 1

wb.save("DS with Python.xlsx")

